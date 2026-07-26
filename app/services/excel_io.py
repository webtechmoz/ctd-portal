"""Build Excel workbooks for reports and project import/export."""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.models.avaliacao import Avaliacao
from app.models.enums import AvaliacaoStatus
from app.models.pilar import Pilar
from app.repositories import pilares as pilar_repo
from app.services.pilar_service import master_out


PROJECTO_HEADERS = [
    "nome",
    "descricao",
    "area",
    "fase",
    "obj_geral",
    "kpis",
    "beneficios",
    "desenvolvedor",
    "orc_moeda",
    "orc_fonte",
    "periodicidade_dias",
    "dias_aberto",
    "proxima_avaliacao",
    "status",
    "responsavel_email",
]

ACTIVIDADE_HEADERS = [
    "projecto_nome",
    "nome",
    "responsavel",
    "prioridade",
    "data_inicio_prevista",
    "data_fim_prevista",
    "status",
]

RUBRICA_HEADERS = ["projecto_nome", "categoria", "valor_alocado", "obs"]
RISCO_HEADERS = ["projecto_nome", "descricao", "probabilidade", "impacto", "mitigacao"]

TEMPLATE_VERSION = "2026.07"


def _wb_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _style_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _avaliacao_metrics(r: Avaliacao) -> dict:
    executado = Decimal("0")
    for o in r.orcamentos or []:
        executado += Decimal(str(o.valor_executado or 0))
    aprovado = Decimal(str(r.pilar.orc_aprovado or 0)) if r.pilar else Decimal("0")
    orc_pct = float((executado / aprovado) * 100) if aprovado else 0.0
    riscos_altos = 0
    risco_map = {x.id: x for x in (r.pilar.riscos if r.pilar else [])}
    for row in r.riscos or []:
        base = risco_map.get(row.risco_id)
        if not base:
            continue
        prob = (
            base.probabilidade.value
            if hasattr(base.probabilidade, "value")
            else str(base.probabilidade or "")
        )
        impacto = (
            base.impacto.value
            if hasattr(base.impacto, "value")
            else str(base.impacto or "")
        )
        if prob == "alta" or impacto == "alto":
            riscos_altos += 1
    st = r.status.value if hasattr(r.status, "value") else (r.status or "submetida")
    return {
        "id": r.id,
        "pilar_id": r.pilar_id,
        "pilar_nome": r.pilar.nome if r.pilar else "",
        "data_sub": r.data_sub.isoformat() if r.data_sub else None,
        "progresso": float(r.progresso or 0),
        "orcamento_pct": round(orc_pct, 1),
        "riscos_altos": riscos_altos,
        "status": st,
        "autor": r.user.name if r.user else None,
        "estado_geral": (r.estado_geral or "")[:180],
    }


def _avaliacoes_query(
    session: Session,
    *,
    pilar_id: int | None = None,
    status: str | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    limit: int | None = None,
):
    stmt = (
        select(Avaliacao)
        .options(
            selectinload(Avaliacao.pilar).selectinload(Pilar.riscos),
            selectinload(Avaliacao.user),
            selectinload(Avaliacao.orcamentos),
            selectinload(Avaliacao.riscos),
        )
        .order_by(Avaliacao.data_sub.desc(), Avaliacao.id.desc())
    )
    if pilar_id:
        stmt = stmt.where(Avaliacao.pilar_id == pilar_id)
    if status:
        try:
            stmt = stmt.where(Avaliacao.status == AvaliacaoStatus(status))
        except ValueError:
            pass
    if date_from:
        stmt = stmt.where(Avaliacao.data_sub >= date_from)
    if date_to:
        stmt = stmt.where(Avaliacao.data_sub <= date_to)
    if limit:
        stmt = stmt.limit(limit)
    return list(session.scalars(stmt).all())


def build_avaliacoes_report(
    session: Session,
    *,
    pilar_id: int | None = None,
    status: str | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
) -> bytes:
    rows = _avaliacoes_query(
        session,
        pilar_id=pilar_id,
        status=status,
        date_from=date_from,
        date_to=date_to,
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Avaliacoes"
    ws.append(
        [
            "id",
            "projecto",
            "data_sub",
            "progresso",
            "orcamento_pct",
            "riscos_altos",
            "status",
            "autor",
            "estado_geral",
        ]
    )
    _style_header(ws)
    for r in rows:
        m = _avaliacao_metrics(r)
        ws.append(
            [
                m["id"],
                m["pilar_nome"],
                m["data_sub"] or "",
                m["progresso"],
                m["orcamento_pct"],
                m["riscos_altos"],
                m["status"],
                m["autor"] or "",
                m["estado_geral"],
            ]
        )
    return _wb_bytes(wb)


def list_avaliacoes_report(
    session: Session,
    *,
    pilar_id: int | None = None,
    status: str | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
) -> list[dict]:
    rows = _avaliacoes_query(
        session,
        pilar_id=pilar_id,
        status=status,
        date_from=date_from,
        date_to=date_to,
        limit=500,
    )
    return [_avaliacao_metrics(r) for r in rows]


def build_import_template(session: Session | None = None) -> bytes:
    from app.catalog_seed import CATALOG_LABELS, CATALOG_SEED
    from app.models.enums import ActividadeStatus, Impacto, PilarStatus, Prioridade, Probabilidade
    from app.services import catalog_service

    wb = Workbook()
    meta = wb.active
    meta.title = "_meta"
    meta.append(["template_version", TEMPLATE_VERSION])
    meta.append(["generated_on", date.today().isoformat()])
    meta.append(
        [
            "notes",
            "Preencha Projecto/Actividades/Rubricas/Riscos. Use valores da folha Listas (codigo).",
        ]
    )

    ws = wb.create_sheet("Projecto")
    ws.append(PROJECTO_HEADERS)
    _style_header(ws)

    wa = wb.create_sheet("Actividades")
    wa.append(ACTIVIDADE_HEADERS)
    _style_header(wa)

    wr = wb.create_sheet("Rubricas")
    wr.append(RUBRICA_HEADERS)
    _style_header(wr)

    wk = wb.create_sheet("Riscos")
    wk.append(RISCO_HEADERS)
    _style_header(wk)

    listas = wb.create_sheet("Listas")
    listas.append(["categoria", "codigo", "etiqueta", "campo_excel"])
    _style_header(listas)

    field_map = {
        "area": "Projecto.area",
        "fase": "Projecto.fase",
        "moeda": "Projecto.orc_moeda",
        "fonte_financiamento": "Projecto.orc_fonte",
    }

    if session is not None:
        for opt in catalog_service.list_all(session):
            if not opt.active:
                continue
            listas.append(
                [opt.category, opt.code, opt.label, field_map.get(opt.category, opt.category)]
            )
    else:
        for cat, rows in CATALOG_SEED.items():
            for code, label in rows:
                listas.append([cat, code, label, field_map.get(cat, cat)])

    for val in PilarStatus:
        listas.append(["status_projecto", val.value, val.value, "Projecto.status"])
    for val in Prioridade:
        listas.append(["prioridade", val.value, val.value, "Actividades.prioridade"])
    for val in ActividadeStatus:
        listas.append(["status_actividade", val.value, val.value, "Actividades.status"])
    for val in Probabilidade:
        listas.append(["probabilidade", val.value, val.value, "Riscos.probabilidade"])
    for val in Impacto:
        listas.append(["impacto", val.value, val.value, "Riscos.impacto"])

    help_ws = wb.create_sheet("Ajuda")
    help_ws.append(["Campo", "Valores aceites"])
    _style_header(help_ws)
    for cat, label in CATALOG_LABELS.items():
        help_ws.append(
            [field_map.get(cat, cat), f"Ver Listas filtrado por categoria={cat} ({label})"]
        )
    help_ws.append(["Projecto.status", "activo | concluido | inactivo"])
    help_ws.append(["Actividades.prioridade", "alta | media | baixa"])
    help_ws.append(["Actividades.status", "activa | cancelada"])
    help_ws.append(["Riscos.probabilidade", "alta | media | baixa"])
    help_ws.append(["Riscos.impacto", "alto | medio | baixo"])
    help_ws.append(["Projecto.responsavel_email", "Email de um utilizador registado"])

    return _wb_bytes(wb)


def export_pilares(session: Session, pilar_ids: list[int] | None = None) -> bytes:
    if pilar_ids:
        pilares = [pilar_repo.get_with_master(session, i) for i in pilar_ids]
        pilares = [p for p in pilares if p]
    else:
        pilares = []
        for p in pilar_repo.list_all(session):
            full = pilar_repo.get_with_master(session, p.id)
            if full:
                pilares.append(full)

    wb = Workbook()
    meta = wb.active
    meta.title = "_meta"
    meta.append(["template_version", TEMPLATE_VERSION])
    meta.append(["exported_on", date.today().isoformat()])
    meta.append(["count", len(pilares)])

    ws = wb.create_sheet("Projecto")
    ws.append(PROJECTO_HEADERS)
    _style_header(ws)
    wa = wb.create_sheet("Actividades")
    wa.append(ACTIVIDADE_HEADERS)
    _style_header(wa)
    wr = wb.create_sheet("Rubricas")
    wr.append(RUBRICA_HEADERS)
    _style_header(wr)
    wk = wb.create_sheet("Riscos")
    wk.append(RISCO_HEADERS)
    _style_header(wk)

    for p in pilares:
        out = master_out(session, p)
        resp_email = out.responsavel_email or ""
        ws.append(
            [
                p.nome,
                p.descricao or "",
                p.area or "",
                p.fase or "",
                p.obj_geral or "",
                p.kpis or "",
                p.beneficios or "",
                p.desenvolvedor or "",
                p.orc_moeda or "MZN",
                p.orc_fonte or "",
                p.periodicidade_dias,
                p.dias_aberto,
                p.proxima_avaliacao.isoformat() if p.proxima_avaliacao else "",
                p.status.value if hasattr(p.status, "value") else str(p.status),
                resp_email,
            ]
        )
        for a in p.actividades:
            wa.append(
                [
                    p.nome,
                    a.nome,
                    a.responsavel or "",
                    a.prioridade.value if hasattr(a.prioridade, "value") else a.prioridade,
                    a.data_inicio_prevista.isoformat() if a.data_inicio_prevista else "",
                    a.data_fim_prevista.isoformat() if a.data_fim_prevista else "",
                    a.status.value if hasattr(a.status, "value") else (a.status or "activa"),
                ]
            )
        for c in p.orcamento_categorias:
            wr.append([p.nome, c.categoria, float(c.valor_alocado or 0), c.obs or ""])
        for r in p.riscos:
            wk.append(
                [
                    p.nome,
                    r.descricao,
                    r.probabilidade.value if hasattr(r.probabilidade, "value") else r.probabilidade,
                    r.impacto.value if hasattr(r.impacto, "value") else r.impacto,
                    r.mitigacao or "",
                ]
            )
    return _wb_bytes(wb)


def _catalog_maps(session: Session) -> dict[str, dict[str, str]]:
    from app.services import catalog_service

    maps: dict[str, dict[str, str]] = {}
    for opt in catalog_service.list_all(session):
        if not opt.active:
            continue
        bucket = maps.setdefault(opt.category, {})
        bucket[opt.code.strip().lower()] = opt.code
        bucket[opt.label.strip().lower()] = opt.code
    return maps


def _resolve_catalog(maps: dict[str, dict[str, str]], category: str, raw: str) -> str | None:
    value = (raw or "").strip()
    if not value:
        return ""
    bucket = maps.get(category) or {}
    return bucket.get(value.lower())


def import_pilares_from_xlsx(
    session: Session, data: bytes, *, dry_run: bool = False
) -> dict:
    from openpyxl import load_workbook

    from app.models.enums import ActividadeStatus, Impacto, PilarStatus, Prioridade, Probabilidade
    from app.models.pilar import (
        Pilar,
        PilarActividade,
        PilarOrcamentoCategoria,
        PilarResponsavel,
        PilarRisco,
    )
    from app.models.user import User
    from app.services.pilar_service import _sync_orc_aprovado

    wb = load_workbook(io.BytesIO(data), data_only=True)
    errors: list[str] = []
    preview: list[dict] = []
    created = updated = 0
    catalog = _catalog_maps(session)

    if "Projecto" not in wb.sheetnames:
        return {
            "ok": False,
            "errors": ["Folha Projecto em falta."],
            "created": 0,
            "updated": 0,
            "preview": [],
            "dry_run": dry_run,
        }

    ws = wb["Projecto"]
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for h in ("nome",):
        if h not in headers:
            errors.append(f"Cabecalho obrigatorio em falta: {h}")
    if errors:
        return {
            "ok": False,
            "errors": errors,
            "created": 0,
            "updated": 0,
            "preview": [],
            "dry_run": dry_run,
        }

    idx = {h: i for i, h in enumerate(headers)}

    def cell(row, key, default=""):
        i = idx.get(key)
        if i is None or i >= len(row):
            return default
        v = row[i]
        return default if v is None else v

    def sheet_index(sheet_name: str):
        sh = wb[sheet_name]
        hdr = [str(c.value or "").strip() for c in next(sh.iter_rows(min_row=1, max_row=1))]
        return list(sh.iter_rows(min_row=2, values_only=True)), {h: i for i, h in enumerate(hdr)}

    acts_by_proj: dict[str, list] = {}
    rubs_by_proj: dict[str, list] = {}
    risks_by_proj: dict[str, list] = {}
    act_idx: dict = {}
    rub_idx: dict = {}
    risk_idx: dict = {}

    if "Actividades" in wb.sheetnames:
        rows_a, act_idx = sheet_index("Actividades")
        for row in rows_a:
            if not row or not any(row):
                continue
            nome_p = str(row[act_idx.get("projecto_nome", 0)] or "").strip()
            if nome_p:
                acts_by_proj.setdefault(nome_p, []).append(row)

    if "Rubricas" in wb.sheetnames:
        rows_r, rub_idx = sheet_index("Rubricas")
        for row in rows_r:
            if not row or not any(row):
                continue
            nome_p = str(row[rub_idx.get("projecto_nome", 0)] or "").strip()
            if nome_p:
                rubs_by_proj.setdefault(nome_p, []).append(row)

    if "Riscos" in wb.sheetnames:
        rows_k, risk_idx = sheet_index("Riscos")
        for row in rows_k:
            if not row or not any(row):
                continue
            nome_p = str(row[risk_idx.get("projecto_nome", 0)] or "").strip()
            if nome_p:
                risks_by_proj.setdefault(nome_p, []).append(row)

    parsed_projects: list[dict] = []

    for row_i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):
            continue
        row_errors: list[str] = []
        nome = str(cell(row, "nome") or "").strip()
        if not nome:
            msg = f"Projecto linha {row_i}: nome obrigatorio."
            errors.append(msg)
            preview.append(
                {"linha": row_i, "nome": "(sem nome)", "action": "erro", "errors": [msg]}
            )
            continue

        existing = session.scalar(select(Pilar).where(Pilar.nome == nome))
        action = "update" if existing else "create"

        area_raw = str(cell(row, "area") or "").strip()
        area = _resolve_catalog(catalog, "area", area_raw)
        if area_raw and area is None:
            row_errors.append(f"area «{area_raw}» nao existe nas Listas de sistema")

        fase_raw = str(cell(row, "fase") or "").strip()
        fase = _resolve_catalog(catalog, "fase", fase_raw)
        if fase_raw and fase is None:
            row_errors.append(f"fase «{fase_raw}» nao existe nas Listas de sistema")

        moeda_raw = str(cell(row, "orc_moeda") or "MZN").strip() or "MZN"
        moeda = _resolve_catalog(catalog, "moeda", moeda_raw)
        if moeda is None:
            row_errors.append(f"orc_moeda «{moeda_raw}» invalida")

        fonte_raw = str(cell(row, "orc_fonte") or "").strip()
        fonte = (
            _resolve_catalog(catalog, "fonte_financiamento", fonte_raw) if fonte_raw else ""
        )
        if fonte_raw and fonte is None:
            row_errors.append(f"orc_fonte «{fonte_raw}» nao existe nas Listas de sistema")

        status_raw = str(cell(row, "status", "activo") or "activo").strip().lower()
        try:
            status = PilarStatus(status_raw)
        except ValueError:
            row_errors.append(f"status «{status_raw}» invalido (activo|concluido|inactivo)")
            status = PilarStatus.activo

        try:
            periodicidade = int(cell(row, "periodicidade_dias", 90) or 90)
            if periodicidade < 1:
                raise ValueError()
        except Exception:
            row_errors.append("periodicidade_dias deve ser um inteiro >= 1")
            periodicidade = 90

        try:
            dias_aberto = int(cell(row, "dias_aberto", 7) or 7)
            if dias_aberto < 0:
                raise ValueError()
        except Exception:
            row_errors.append("dias_aberto deve ser um inteiro >= 0")
            dias_aberto = 7

        email = str(cell(row, "responsavel_email") or "").strip().lower()
        resp_user = None
        if email:
            resp_user = session.scalar(select(User).where(User.email == email))
            if not resp_user:
                row_errors.append(f"responsavel_email «{email}» nao encontrado")

        for i, arow in enumerate(acts_by_proj.get(nome, []), start=1):
            anome = str(arow[act_idx.get("nome", 1)] or "").strip()
            if not anome:
                row_errors.append(f"Actividade #{i}: nome obrigatorio")
                continue
            prio = str(arow[act_idx.get("prioridade", 3)] or "media").strip().lower()
            if prio not in Prioridade.__members__:
                row_errors.append(f"Actividade «{anome}»: prioridade «{prio}» invalida")
            st = str(arow[act_idx.get("status", 6)] or "activa").strip().lower()
            if st not in ActividadeStatus.__members__:
                row_errors.append(f"Actividade «{anome}»: status «{st}» invalido")

        for i, rrow in enumerate(rubs_by_proj.get(nome, []), start=1):
            cat = str(rrow[rub_idx.get("categoria", 1)] or "").strip()
            if not cat:
                row_errors.append(f"Rubrica #{i}: categoria obrigatoria")
                continue
            val = rrow[rub_idx.get("valor_alocado", 2)]
            try:
                Decimal(str(val if val is not None else 0))
            except Exception:
                row_errors.append(f"Rubrica «{cat}»: valor_alocado invalido")

        for i, krow in enumerate(risks_by_proj.get(nome, []), start=1):
            desc = str(krow[risk_idx.get("descricao", 1)] or "").strip()
            if not desc:
                row_errors.append(f"Risco #{i}: descricao obrigatoria")
                continue
            prob = str(krow[risk_idx.get("probabilidade", 2)] or "media").strip().lower()
            if prob not in Probabilidade.__members__:
                row_errors.append(f"Risco «{desc[:40]}»: probabilidade «{prob}» invalida")
            imp = str(krow[risk_idx.get("impacto", 3)] or "medio").strip().lower()
            if imp not in Impacto.__members__:
                row_errors.append(f"Risco «{desc[:40]}»: impacto «{imp}» invalido")

        if row_errors:
            for e in row_errors:
                errors.append(f"Projecto «{nome}»: {e}")
        elif action == "create":
            created += 1
        else:
            updated += 1

        preview.append(
            {
                "linha": row_i,
                "nome": nome,
                "action": action if not row_errors else "erro",
                "area": area or area_raw,
                "fase": fase or fase_raw,
                "status": status.value,
                "actividades": len(acts_by_proj.get(nome, [])),
                "rubricas": len(rubs_by_proj.get(nome, [])),
                "riscos": len(risks_by_proj.get(nome, [])),
                "errors": row_errors,
            }
        )

        if not row_errors:
            parsed_projects.append(
                {
                    "nome": nome,
                    "existing": existing,
                    "descricao": str(cell(row, "descricao") or ""),
                    "area": area or "",
                    "fase": fase or "",
                    "obj_geral": str(cell(row, "obj_geral") or ""),
                    "kpis": str(cell(row, "kpis") or ""),
                    "beneficios": str(cell(row, "beneficios") or ""),
                    "desenvolvedor": str(cell(row, "desenvolvedor") or "") or None,
                    "orc_moeda": moeda or "MZN",
                    "orc_fonte": fonte or None,
                    "periodicidade_dias": periodicidade,
                    "dias_aberto": dias_aberto,
                    "status": status,
                    "resp_user": resp_user,
                }
            )

    result_base = {
        "ok": len(errors) == 0,
        "errors": errors,
        "created": created,
        "updated": updated,
        "preview": preview,
        "dry_run": dry_run,
    }

    if dry_run or errors:
        return result_base

    for item in parsed_projects:
        existing = item["existing"]
        if existing:
            pilar = existing
        else:
            pilar = Pilar(nome=item["nome"])
            session.add(pilar)
        pilar.descricao = item["descricao"]
        pilar.area = item["area"]
        pilar.fase = item["fase"]
        pilar.obj_geral = item["obj_geral"]
        pilar.kpis = item["kpis"]
        pilar.beneficios = item["beneficios"]
        pilar.desenvolvedor = item["desenvolvedor"]
        pilar.orc_moeda = item["orc_moeda"]
        pilar.orc_fonte = item["orc_fonte"]
        pilar.periodicidade_dias = item["periodicidade_dias"]
        pilar.dias_aberto = item["dias_aberto"]
        pilar.status = item["status"]
        session.flush()

        if item["resp_user"]:
            pilar.responsaveis.clear()
            session.flush()
            pilar.responsaveis.append(PilarResponsavel(user_id=item["resp_user"].id))

        nome = item["nome"]
        if nome in acts_by_proj:
            pilar.actividades.clear()
            session.flush()
            for i, arow in enumerate(acts_by_proj[nome]):
                anome = str(arow[act_idx.get("nome", 1)] or "").strip()
                if not anome:
                    continue
                prio = str(arow[act_idx.get("prioridade", 3)] or "media").strip().lower()
                st = str(arow[act_idx.get("status", 6)] or "activa").strip().lower()
                pilar.actividades.append(
                    PilarActividade(
                        nome=anome,
                        responsavel=str(arow[act_idx.get("responsavel", 2)] or ""),
                        prioridade=Prioridade(prio),
                        status=ActividadeStatus(st),
                        ordem=i,
                    )
                )

        if nome in rubs_by_proj:
            pilar.orcamento_categorias.clear()
            session.flush()
            for i, rrow in enumerate(rubs_by_proj[nome]):
                cat = str(rrow[rub_idx.get("categoria", 1)] or "").strip()
                if not cat:
                    continue
                val = rrow[rub_idx.get("valor_alocado", 2)] or 0
                pilar.orcamento_categorias.append(
                    PilarOrcamentoCategoria(
                        categoria=cat,
                        valor_alocado=Decimal(str(val)),
                        obs=str(rrow[rub_idx.get("obs", 3)] or "") or None,
                        ordem=i,
                    )
                )
            _sync_orc_aprovado(pilar)

        if nome in risks_by_proj:
            pilar.riscos.clear()
            session.flush()
            for i, krow in enumerate(risks_by_proj[nome]):
                desc = str(krow[risk_idx.get("descricao", 1)] or "").strip()
                if not desc:
                    continue
                prob = str(krow[risk_idx.get("probabilidade", 2)] or "media").strip().lower()
                imp = str(krow[risk_idx.get("impacto", 3)] or "medio").strip().lower()
                pilar.riscos.append(
                    PilarRisco(
                        descricao=desc,
                        probabilidade=Probabilidade(prob),
                        impacto=Impacto(imp),
                        mitigacao=str(krow[risk_idx.get("mitigacao", 4)] or "") or None,
                        ordem=i,
                    )
                )

    session.flush()
    return {**result_base, "dry_run": False}
